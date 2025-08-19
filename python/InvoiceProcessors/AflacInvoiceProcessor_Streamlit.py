import streamlit as st
import pandas as pd

st.title("Aflac Invoice and Support Generator")

# Upload both files
invoice_file = st.file_uploader("Upload Invoice Excel File", type=["xlsx"], key="invoice")
template_file = st.file_uploader("Upload Template Excel File (with Code Map sheet)", type=["xlsx"], key="template")

if invoice_file and template_file:
    # Load invoice data
    invoice_xls = pd.ExcelFile(invoice_file, engine='openpyxl')
    detail_df = pd.read_excel(invoice_xls, sheet_name='Detail', engine='openpyxl')

    # Load template data and detect first sheet dynamically
    template_xls = pd.ExcelFile(template_file, engine='openpyxl')
    first_sheet_name = template_xls.sheet_names[0]
    template_df = pd.read_excel(template_xls, sheet_name=first_sheet_name, engine='openpyxl')

    # Load Code Map sheet
    code_map_df = pd.read_excel(template_xls, sheet_name='Code Map', engine='openpyxl')

    # Table 1: Summary by Company (grouped properly to avoid duplicates)
    premium_summary = detail_df.groupby('Company')['Monthly Premium'].sum().reset_index()
    premium_summary.columns = ['Row Labels', 'Sum of Monthly Premium']

    result_df = premium_summary.merge(
        code_map_df[['Invoice Company Code', 'Company Description']],
        left_on='Row Labels', right_on='Invoice Company Code', how='left'
    ).drop_duplicates(subset=['Row Labels'])

    result_df.rename(columns={'Company Description': 'Full Company Name'}, inplace=True)
    final_df = result_df[['Row Labels', 'Sum of Monthly Premium', 'Full Company Name']]

    # Table 2: Hierarchical breakdown by Company and Division (filtered)
    code_map_filtered = code_map_df[code_map_df['Division Description'].notna()]
    valid_companies = code_map_filtered['Invoice Company Code'].unique()
    filtered_detail_df = detail_df[detail_df['Company'].isin(valid_companies)]

    breakdown_rows = []
    company_totals = filtered_detail_df.groupby('Company')['Monthly Premium'].sum().reset_index()
    for company, comp_premium in company_totals.values:
        breakdown_rows.append({'Label': company, 'Monthly Premium': comp_premium})
        divisions = filtered_detail_df[filtered_detail_df['Company'] == company].groupby('Division')['Monthly Premium'].sum().reset_index()
        for div, div_premium in divisions.values:
            if pd.notna(div):
                breakdown_rows.append({'Label': f"  {div}", 'Monthly Premium': div_premium})
    breakdown_df = pd.DataFrame(breakdown_rows)

    # Table 3: THC & HHI breakdown using Department column
    thchhi_df = detail_df[detail_df['Company'].isin(['THC', 'HHI'])]
    total_thchhi = thchhi_df['Monthly Premium'].sum()

    department_summary = thchhi_df.groupby('Department')['Monthly Premium'].sum().reset_index()

    department_rows = [{'Department': 'THC & HHI', 'Sum of Monthly Premium': total_thchhi}]
    cc_desc_map = template_df[['CC', 'DESC']].dropna()
    cc_desc_dict = dict(zip(cc_desc_map['CC'].astype(str), cc_desc_map['DESC']))

    for dept_code, premium in department_summary.values:
        if pd.notna(dept_code) and isinstance(dept_code, str) and len(dept_code) >= 2:
            stripped_code = dept_code[2:]
            department_name = cc_desc_dict.get(stripped_code, stripped_code)
            department_rows.append({'Department': department_name, 'Sum of Monthly Premium': premium})

    department_df = pd.DataFrame(department_rows)

    # Display tables in Streamlit
    st.subheader("Summary Table")
    st.dataframe(final_df)

    st.subheader("Company & Division Breakdown (Filtered)")
    st.dataframe(breakdown_df)

    st.subheader("THC & HHI Department Breakdown")
    st.dataframe(department_df)

    # Export to Excel
    def convert_df_to_excel(df1, df2, df3):
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Aflac Invoice and Support'

        # Write Table 1 at A1
        for r in dataframe_to_rows(df1, index=False, header=True):
            ws.append(r)

        # Write Table 2 at E1
        for i, r in enumerate(dataframe_to_rows(df2, index=False, header=True), start=1):
            for j, val in enumerate(r, start=5):  # Column E is index 5
                ws.cell(row=i, column=j, value=val)

        # Write Table 3 at H1
        for i, r in enumerate(dataframe_to_rows(df3, index=False, header=True), start=1):
            for j, val in enumerate(r, start=8):  # Column H is index 8
                ws.cell(row=i, column=j, value=val)

        wb.save(output)
        return output.getvalue()

    excel_data = convert_df_to_excel(final_df, breakdown_df, department_df)
    st.download_button(
        label="Download Aflac Invoice and Support Excel",
        data=excel_data,
        file_name="Aflac_Invoice_and_Support.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
