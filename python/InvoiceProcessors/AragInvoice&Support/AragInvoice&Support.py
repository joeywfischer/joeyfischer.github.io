import streamlit as st
import pandas as pd

st.title("Arag Invoice and Support Generator")

# Upload both files
invoice_file = st.file_uploader("Upload Invoice Excel File", type=["xlsx"], key="invoice")
template_file = st.file_uploader("Upload Medius Template Excel File (with Code Map sheet)", type=["xlsx"], key="template")

if invoice_file and template_file:
    # Load invoice data
    invoice_xls = pd.ExcelFile(invoice_file, engine='openpyxl')
    detail_df = pd.read_excel(invoice_xls, sheet_name='ARAG Self Bill Detail', engine='openpyxl')

    # Load template data and detect first sheet dynamically
    template_xls = pd.ExcelFile(template_file, engine='openpyxl')
    first_sheet_name = template_xls.sheet_names[0]
    template_df = pd.read_excel(template_xls, sheet_name=first_sheet_name, engine='openpyxl')

    # Load Code Map sheet
    code_map_df = pd.read_excel(template_xls, sheet_name='Code Map', engine='openpyxl')

    # Table 1: Summary by Company
    premium_summary = detail_df.groupby('Company')['Real Premium'].sum().reset_index()
    premium_summary.columns = ['Row Labels', 'Sum of Real Premium']

    result_df = premium_summary.merge(
        code_map_df[['Invoice Company Code', 'Company Description']],
        left_on='Row Labels', right_on='Invoice Company Code', how='left'
    ).drop_duplicates(subset=['Row Labels'])

    result_df.rename(columns={'Company Description': 'Full Company Name'}, inplace=True)
    final_df = result_df[['Row Labels', 'Sum of Real Premium', 'Full Company Name']]

    # Add Grand Total row
    grand_total = final_df['Sum of Real Premium'].sum()
    total_row = pd.DataFrame([{
        'Row Labels': 'Grand Total',
        'Sum of Real Premium': grand_total,
        'Full Company Name': ''
    }])
    final_df = pd.concat([final_df, total_row], ignore_index=True)

    # Table 2: Hierarchical breakdown using Division Descriptions
    code_map_filtered = code_map_df[code_map_df['Division Description'].notna()]
    valid_companies = code_map_filtered['Invoice Company Code'].unique()
    filtered_detail_df = detail_df[detail_df['Company'].isin(valid_companies)]

    breakdown_rows = []
    company_totals = filtered_detail_df.groupby('Company')['Real Premium'].sum().reset_index()
    for company, comp_premium in company_totals.values:
        breakdown_rows.append({'Label': company, 'Real Premium': comp_premium})
        company_divisions = filtered_detail_df[filtered_detail_df['Company'] == company]
        company_divisions = company_divisions.merge(
            code_map_df[['Division Code', 'Division Description']],
            left_on='Division', right_on='Division Code', how='left'
        )
        division_totals = company_divisions.groupby('Division Description')['Real Premium'].sum().reset_index()
        for div_desc, div_premium in division_totals.values:
            if pd.notna(div_desc):
                breakdown_rows.append({'Label': f"  {div_desc}", 'Real Premium': div_premium})
    breakdown_df = pd.DataFrame(breakdown_rows)

    # Table 3: THC & HHI breakdown using Department column
    thchhi_df = detail_df[detail_df['Company'].isin(['THC', 'HHI'])]
    total_thchhi = thchhi_df['Real Premium'].sum()

    department_summary = thchhi_df.groupby('Department')['Real Premium'].sum().reset_index()
    department_summary['Stripped Code'] = department_summary['Department'].apply(
        lambda x: str(x)[2:] if pd.notna(x) and isinstance(x, (str, int, float)) and len(str(x)) >= 2 else str(x)
    )

    cc_desc_map = template_df[['CC', 'DESC']].dropna()
    cc_desc_map['CC'] = cc_desc_map['CC'].apply(lambda x: str(int(x)) if isinstance(x, float) else str(x))
    cc_desc_dict = dict(zip(cc_desc_map['CC'], cc_desc_map['DESC']))

    department_summary['Mapped DESC'] = department_summary['Stripped Code'].apply(
        lambda x: cc_desc_dict.get(x, x)
    )

    department_rows = [{'Department': 'THC & HHI', 'Sum of Real Premium': total_thchhi}]
    for _, row in department_summary.iterrows():
        department_rows.append({
            'Department': row['Mapped DESC'],
            'Sum of Real Premium': row['Real Premium']
        })
    department_df = pd.DataFrame(department_rows)

    # Export to Excel with formatting
    def convert_df_to_excel(df1, df2, df3):
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, NamedStyle
        from openpyxl.utils.dataframe import dataframe_to_rows

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Arag Invoice and Support'

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        bold_font = Font(bold=True)
        accounting_style = NamedStyle(name="accounting_style", number_format='_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)')

        # Helper to auto-adjust column widths
        def adjust_column_widths(ws, start_col, df):
            for i, col in enumerate(df.columns, start=start_col):
                max_length = max([len(str(cell)) for cell in [col] + df[col].astype(str).tolist()])
                col_letter = ws.cell(row=1, column=i).column_letter
                ws.column_dimensions[col_letter].width = max_length + 2

        # Table 1 at A1
        grand_total_row_idx = None
        for r_idx, r in enumerate(dataframe_to_rows(df1, index=False, header=True), start=1):
            for c_idx, val in enumerate(r, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                elif c_idx == 2 and r_idx > 1:
                    cell.style = accounting_style

                # Detect the "Grand Total" row (first column equals "Grand Total")
                if r_idx > 1 and c_idx == 1 and str(val).strip() == 'Grand Total':
                    grand_total_row_idx = r_idx

                # Bold the entire "Grand Total" row
                if grand_total_row_idx is not None and r_idx == grand_total_row_idx:
                    cell.font = bold_font
        adjust_column_widths(ws, 1, df1)

        # Table 2 at E1
        for r_idx, r in enumerate(dataframe_to_rows(df2, index=False, header=True), start=1):
            for c_idx, val in enumerate(r, start=5):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                elif c_idx == 6 and r_idx > 1:
                    cell.style = accounting_style
                if c_idx == 5 and r_idx > 1 and not str(val).startswith("  "):
                    cell.font = bold_font
        adjust_column_widths(ws, 5, df2)

        # Table 3 at H1
        for r_idx, r in enumerate(dataframe_to_rows(df3, index=False, header=True), start=1):
            for c_idx, val in enumerate(r, start=8):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                elif c_idx == 9 and r_idx > 1:
                    cell.style = accounting_style
                if c_idx == 8 and r_idx > 1 and str(val).strip() == 'THC & HHI':
                    cell.font = bold_font
        adjust_column_widths(ws, 8, df3)

        wb.save(output)
        return output.getvalue()

    excel_data = convert_df_to_excel(final_df, breakdown_df, department_df)
    st.download_button(
        label="Download Arag Invoice and Support Excel",
        data=excel_data,
        file_name="Arag_Invoice_and_Support.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
