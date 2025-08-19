import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

st.title("Aflac Invoice Processor")

invoice_file = st.file_uploader("Upload Aflac Invoice Excel File", type=["xlsx"])
template_file = st.file_uploader("Upload Medius Template Excel File", type=["xlsx"])

if invoice_file and template_file:
    try:
        # Load invoice and template data
        df_invoice = pd.read_excel(invoice_file, sheet_name='Detail', engine='openpyxl')
        df_template = pd.read_excel(template_file, sheet_name=0, engine='openpyxl')
        df_code_map = pd.read_excel(template_file, sheet_name='Code Map', engine='openpyxl')

        # Normalize key columns
        df_invoice['Company'] = df_invoice['Company'].astype(str).str.strip().str.upper()
        df_invoice['Division'] = df_invoice['Division'].astype(str).str.strip()
        df_invoice['Monthly Premium'] = pd.to_numeric(df_invoice['Monthly Premium'], errors='coerce')

        # [Mapping logic remains unchanged — omitted for brevity]

        # --- Medius Template Output ---
        output_template = io.BytesIO()
        df_template.to_excel(output_template, index=False, sheet_name='Updated Medius Template', engine='openpyxl')
        output_template.seek(0)

        # --- Invoice Support Table ---
        df_support = df_invoice[['Company', 'Monthly Premium']].dropna().groupby('Company')['Monthly Premium'].sum().reset_index()
        df_support = df_support.rename(columns={'Company': 'Invoice Company Code', 'Monthly Premium': 'Sum of Monthly Premium'})
        df_support['Full Company Name'] = df_support['Invoice Company Code'].map(
            dict(zip(
                df_code_map['Invoice Company Code'].astype(str).str.upper(),
                df_code_map['Company Description'].astype(str).str.strip()
            ))
        )
        df_support = df_support.dropna()
        df_support = df_support[['Invoice Company Code', 'Sum of Monthly Premium', 'Full Company Name']]

        # --- Additional Break-Down Table ---
        valid_divisions = df_code_map['Division Code'].dropna().astype(str).str.strip().unique()
        df_division_filtered = df_invoice[df_invoice['Division'].isin(valid_divisions)]
        df_division_summary = df_division_filtered.groupby('Division')['Monthly Premium'].sum().reset_index()
        df_division_summary = df_division_summary.rename(columns={'Division': 'Division Code', 'Monthly Premium': 'Sum of Monthly Premium'})
        df_division_summary['Division Description'] = df_division_summary['Division Code'].map(
            dict(zip(
                df_code_map['Division Code'].astype(str).str.strip(),
                df_code_map['Division Description'].astype(str).str.strip()
            ))
        )
        df_division_summary = df_division_summary.dropna()
        df_division_summary = df_division_summary[['Division Description', 'Sum of Monthly Premium']]

        # --- Save with formatting ---
        output_support = io.BytesIO()
        with pd.ExcelWriter(output_support, engine='openpyxl') as writer:
            # Write all sheets
            df_template.to_excel(writer, index=False, sheet_name='Updated Medius Template')
            df_support.to_excel(writer, index=False, sheet_name='Invoice Support')
            df_division_summary.to_excel(writer, index=False, sheet_name='Additional Break-Down')

            for sheet_name in ['Invoice Support', 'Additional Break-Down']:
                worksheet = writer.sheets[sheet_name]
                header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                header_font = Font(bold=True)

                # Format header
                for col_num in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font

                # Adjust column widths
                for col in worksheet.columns:
                    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                    worksheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

                # Format currency column (column 2)
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)
                    cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                # Add and style Grand Total row
                total = sum([worksheet.cell(row=r, column=2).value or 0 for r in range(2, worksheet.max_row + 1)])
                total_row = worksheet.max_row + 1
                worksheet.cell(row=total_row, column=1).value = 'Grand Total'
                worksheet.cell(row=total_row, column=1).font = header_font
                worksheet.cell(row=total_row, column=1).fill = header_fill
                worksheet.cell(row=total_row, column=2).value = total
                worksheet.cell(row=total_row, column=2).number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

        output_support.seek(0)

        # --- Streamlit Outputs ---
        st.success("Processing complete!")

        st.download_button(
            label="Download Updated Medius Template",
            data=output_template,
            file_name="Complete_Aflac_Medius_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.download_button(
            label="Download Aflac Invoice and Support",
            data=output_support,
            file_name="Aflac_Invoice_and_Support.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"An error occurred: {e}")


